import poplib
import email
from email.parser import Parser
from email.policy import default
import time
import os
import logging
from datetime import datetime, timedelta
import re
from openpyxl import Workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.header import decode_header
from email.utils import parsedate_to_datetime
import sqlite3
import csv
from config_manager import ConfigManager

import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
# 初始化配置管理器
config_manager = ConfigManager()
config = config_manager.get_all_configs()



# 添加短信通知所需的模块
import ssl
import urllib.request
import urllib.parse
import urllib.error
from xml.dom.minidom import parseString

# 配置参数
email_address = config['email']['export_email']
password = config['email']['export_password']
pop3_server = config['email']['pop3_server']
pop3_port = config['email']['pop3_port']

# SMTP配置（用于发送回复邮件）
smtp_server = config['email']['smtp_server']
smtp_port = config['email']['smtp_port']

# 关键词配置
keywords = config['keywords']['export']

# 关键词 -> 中文货名映射（由配置文件自动维护；未知关键词默认回填英文关键词）
keyword_translation = config_manager.get_keyword_translation_map()

def get_chinese_goods_name(main_keyword: str, fallback_english: str) -> str:
    """根据关键词获取中文货名（配置缺失时使用英文兜底）"""
    try:
        if not main_keyword:
            return fallback_english
        return keyword_translation.get(main_keyword, fallback_english)
    except Exception:
        return fallback_english
# SQLite数据库文件（只记录匹配关键词并已回复的邮件）
db_file = config['files']['export_db']

# 主日志文件（记录所有邮件的处理状态，按配置自动清理）
LOG_CSV_FILE = config['files']['export_log']

# 按需求：
# - 自动检测只检测最近 50 天的邮件
# - 日志文件每 51 天清理一次（避免日志过大）
SCAN_DAYS = 50
try:
    LOG_RETENTION_DAYS = int(config['settings'].get('log_retention_days', 51))
except Exception:
    LOG_RETENTION_DAYS = 51

# 短信配置
SMS_ACCOUNT = config['sms']['account']
SMS_PASSWORD = config['sms']['password']
SMS_MOBILES = config['sms']['mobiles']
SMS_CONTENT_TEMPLATE = config['sms']['export_template']
SMS_API_URL = config['sms']['api_url']

# 设置日志
logging.basicConfig(
    level=logging.DEBUG,  # 改为DEBUG级别
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()  # 只输出到控制台，不保存到文件
    ]
)
###在导入部分添加的功能

try:
    from statistics_system import StatisticsSystem
    STATS_SYSTEM_AVAILABLE = True
except ImportError:
    STATS_SYSTEM_AVAILABLE = False
    logging.warning("⚠️ 统计系统模块不可用，统计功能将受限")



def init_log_file():
    """初始化或清理日志文件，只保留 LOG_RETENTION_DAYS 天内的记录"""
    try:
        # 如果日志文件不存在，创建它并写入标题行
        if not os.path.exists(LOG_CSV_FILE):
            with open(LOG_CSV_FILE, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['timestamp', 'email_uid', 'sender', 'subject', 
                                'has_keyword', 'excel_sent', 'matched_keywords', 'container_count'])
            logging.info(f"✅ 创建日志文件: {LOG_CSV_FILE}")
            return True
        
        # 清理历史日志记录
        cleanup_old_log_entries()
        return True
    except Exception as e:
        logging.error(f"❌ 初始化日志文件失败: {e}")
        return False

def cleanup_old_log_entries():
    """清理超过 LOG_RETENTION_DAYS 天的日志记录"""
    try:
        cutoff_time = datetime.now() - timedelta(days=LOG_RETENTION_DAYS)
        
        # 读取所有记录
        rows = []
        with open(LOG_CSV_FILE, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader)  # 读取标题行
            rows.append(header)
            
            for row in reader:
                if len(row) >= 1:
                    timestamp_str = row[0]
                    try:
                        # 解析时间戳
                        log_time = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                        # 只保留 LOG_RETENTION_DAYS 天内的记录
                        if log_time >= cutoff_time:
                            rows.append(row)
                    except:
                        # 如果时间戳格式错误，保留该行
                        rows.append(row)
        
        # 写回文件
        with open(LOG_CSV_FILE, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        
        logging.info(f"🗑️ 已清理超过 {LOG_RETENTION_DAYS} 天的日志记录，当前保留 {len(rows)-1} 条记录")
        return True
    except Exception as e:
        logging.error(f"❌ 清理日志记录失败: {e}")
        return False

def is_email_processed(email_uid):
    """检查邮件是否已在日志中处理过"""
    try:
        if not os.path.exists(LOG_CSV_FILE):
            return False
        
        with open(LOG_CSV_FILE, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)  # 跳过标题行
            
            for row in reader:
                if len(row) >= 2 and row[1] == email_uid:
                    return True
        
        return False
    except Exception as e:
        logging.error(f"❌ 检查邮件处理状态失败: {e}")
        return False

def log_email_processed(email_uid, sender, subject, has_keyword=False, excel_sent=0, matched_keywords="", container_count=0):
    """记录邮件处理状态到日志文件"""
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 截断过长的字段
        sender_display = sender[:100] if len(sender) > 100 else sender
        subject_display = subject[:200] if len(subject) > 200 else subject
        matched_keywords_display = matched_keywords[:100] if len(matched_keywords) > 100 else matched_keywords
        
        with open(LOG_CSV_FILE, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([timestamp, email_uid, sender_display, subject_display, 
                           int(has_keyword), excel_sent, matched_keywords_display, container_count])
        
        logging.info(f"📝 已记录邮件处理状态: {email_uid}")
        return True
    except Exception as e:
        logging.error(f"❌ 记录邮件处理状态失败: {e}")
        return False

def init_database():
    """初始化数据库 - 只保存匹配到关键词且已发送Excel的邮件"""
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        
        # 创建已处理邮件表 - 添加中英文货名字段
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS keyword_emails (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email_uid TEXT NOT NULL,
            sender TEXT NOT NULL,
            sender_address TEXT,
            subject TEXT NOT NULL,
            received_date TEXT,
            processed_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            matched_keywords TEXT NOT NULL,
            excel_sent INTEGER DEFAULT 1,
            txt_attachment TEXT,
            container_count INTEGER DEFAULT 0,
            attachment_names TEXT,
            english_goods_descriptions TEXT,  -- 新增：英文货名列表
            chinese_goods_descriptions TEXT,  -- 新增：中文货名列表
            sync_source TEXT DEFAULT '',  -- 新增：同步来源
            UNIQUE(email_uid)
        )
        ''')
        
        # 检查并添加缺失的列（如果表已存在但缺少新列）
        cursor.execute("PRAGMA table_info(keyword_emails)")
        columns = [column[1] for column in cursor.fetchall()]
        
        # 添加缺失的列
        if 'english_goods_descriptions' not in columns:
            cursor.execute('ALTER TABLE keyword_emails ADD COLUMN english_goods_descriptions TEXT')
            logging.info("🔄 已添加缺失的列: english_goods_descriptions")
        
        if 'chinese_goods_descriptions' not in columns:
            cursor.execute('ALTER TABLE keyword_emails ADD COLUMN chinese_goods_descriptions TEXT')
            logging.info("🔄 已添加缺失的列: chinese_goods_descriptions")
        
        if 'sync_source' not in columns:
            cursor.execute('ALTER TABLE keyword_emails ADD COLUMN sync_source TEXT DEFAULT ""')
            logging.info("🔄 已添加缺失的列: sync_source")
        
        # 创建索引以提高查询速度
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_email_uid ON keyword_emails(email_uid)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_processed_date ON keyword_emails(processed_date)')
        
        # 清理旧记录（保留最近90天的记录）
        ninety_days_ago = datetime.now() - timedelta(days=90)
        cursor.execute('DELETE FROM keyword_emails WHERE processed_date < ?', (ninety_days_ago.strftime('%Y-%m-%d'),))
        deleted_count = cursor.rowcount
        
        conn.commit()
        conn.close()
        
        if deleted_count > 0:
            logging.info(f"🗑️ 已清理 {deleted_count} 条旧记录（90天前）")
        
        logging.info("✅ 数据库初始化完成")
        return True
    except Exception as e:
        logging.error(f"❌ 数据库初始化失败: {e}")
        return False

def save_keyword_email(email_uid, sender, sender_address, subject, received_date, matched_keywords, 
                       txt_attachment=None, container_count=0, attachment_names="", 
                       english_goods_descriptions="", chinese_goods_descriptions=""):
    """保存匹配到关键词且已发送Excel的邮件信息到数据库"""
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        
        # 插入新记录
        cursor.execute('''
        INSERT OR REPLACE INTO keyword_emails 
        (email_uid, sender, sender_address, subject, received_date, matched_keywords, 
         txt_attachment, container_count, attachment_names, 
         english_goods_descriptions, chinese_goods_descriptions)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (email_uid, sender, sender_address, subject, received_date, matched_keywords, 
              txt_attachment, container_count, attachment_names,
              english_goods_descriptions, chinese_goods_descriptions))
        
        conn.commit()
        conn.close()
        logging.info(f"✅ 关键词邮件已保存到数据库: {subject}")
        return True
    except Exception as e:
        logging.error(f"❌ 保存关键词邮件失败: {e}")
        return False



#新加入函数20251215
def add_attachment_statistics(email_uid, sender, sender_address, subject, received_date, 
                             txt_attachment, container_data, matched_keywords):
    """添加附件统计记录"""
    try:
        if not STATS_SYSTEM_AVAILABLE:
            logging.debug("统计系统不可用，跳过统计记录")
            return False
        
        stats = StatisticsSystem()
        
        if txt_attachment and container_data:
            # 检查是否包含危险品
            has_dangerous = 1 if container_data and len(container_data) > 0 else 0
            
            attachment_info = {
                'attachment_name': txt_attachment,
                'process_date': datetime.now().strftime('%Y-%m-%d'),
                'has_dangerous': has_dangerous,
                'matched_keywords': matched_keywords if matched_keywords else '出口舱单匹配',
                'sender_email': sender_address,
                'subject': subject
            }
            
            # 添加记录
            success = stats.add_attachment_record('export', attachment_info)
            if success:
                logging.info(f"✅ 已添加出口附件统计记录: {txt_attachment}")
            return success
        return False
    except Exception as e:
        logging.error(f"❌ 添加附件统计失败: {e}")
        return False

def decode_email_header(header):
    """解码邮件头，处理MIME编码的文本"""
    if not header:
        return ""
    
    try:
        # 如果header是bytes类型，先解码为字符串
        if isinstance(header, bytes):
            header = header.decode('utf-8', errors='ignore')
        
        # 检查是否为MIME编码格式
        if '=?' in header and '?=' in header:
            decoded_parts = []
            for part, charset in decode_header(header):
                if isinstance(part, bytes):
                    # 如果是bytes类型，使用指定字符集解码
                    if charset:
                        try:
                            decoded_parts.append(part.decode(charset, errors='ignore'))
                        except:
                            decoded_parts.append(part.decode('utf-8', errors='ignore'))
                    else:
                        decoded_parts.append(part.decode('utf-8', errors='ignore'))
                else:
                    # 如果不是bytes类型，直接添加
                    decoded_parts.append(str(part))
            return ''.join(decoded_parts)
        else:
            # 如果不是MIME编码格式，直接返回
            return header
    except Exception as e:
        logging.warning(f"解码邮件头失败: {e}, 原始内容: {header}")
        return header

def extract_email_address(email_string):
    """从邮件字符串中提取邮箱地址"""
    try:
        # 解码邮件头
        decoded_string = decode_email_header(email_string)
        
        # 使用正则表达式提取邮箱地址
        # 匹配 <邮箱地址> 格式
        match = re.search(r'<([^>]+)>', decoded_string)
        if match:
            return match.group(1).strip()
        
        # 如果没有尖括号，直接返回解码后的字符串
        # 检查是否为有效的邮箱格式
        if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', decoded_string.strip()):
            return decoded_string.strip()
        
        # 如果不是标准格式，尝试其他方式提取
        # 查找包含@的字符串
        email_match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', decoded_string)
        if email_match:
            return email_match.group(0).strip()
        
        return decoded_string.strip()
    except Exception as e:
        logging.error(f"提取邮箱地址失败: {e}, 原始字符串: {email_string}")
        return email_string

def normalize_keyword(keyword):
    """标准化关键词（大写并移除空格）"""
    return keyword.upper().replace(' ', '')

def check_keywords_in_text(text):
    """检查文本中是否包含关键词（标准化比较）"""
    if not text:
        return []
    
    # 标准化文本（大写并移除空格）
    normalized_text = text.upper().replace(' ', '')
    
    found_keywords = []
    for keyword in keywords:
        # 标准化关键词
        normalized_keyword = normalize_keyword(keyword)
        if normalized_keyword in normalized_text:
            found_keywords.append(keyword)
    
    return found_keywords

def get_email_body(msg):
    """提取邮件正文内容"""
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition"))
            
            # 跳过附件，只处理正文
            if "attachment" not in content_disposition:
                if content_type == "text/plain":
                    # 解码文本部分
                    try:
                        body += part.get_payload(decode=True).decode('utf-8', errors='ignore')
                    except:
                        pass
    else:
        # 如果不是多部分邮件，直接获取内容
        content_type = msg.get_content_type()
        if content_type == "text/plain":
            try:
                body = msg.get_payload(decode=True).decode('utf-8', errors='ignore')
            except:
                pass
    return body

def is_export_manifest(txt_content):
    """判断TXT内容是否为出口舱单"""
    try:
        if not txt_content:
            return False
        
        # 检查前500个字符
        sample = txt_content[:500] if len(txt_content) > 500 else txt_content
        
        # 特征1: 出口舱单通常以"00NCLCONTAINER LIST"开头
        if "00NCLCONTAINER LIST" in sample:
            logging.info("✅ 检测到出口舱单格式: 以00NCLCONTAINER LIST开头")
            return True
        
        # 特征2: 进口舱单通常以"00:IFCSUM:"开头或有冒号分隔格式
        if "00:IFCSUM:" in sample:
            logging.info("❌ 检测到进口舱单格式: 以00:IFCSUM开头")
            return False
        
        # 特征3: 检查是否有冒号分隔的格式（进口舱单特征）
        lines = txt_content.split('\n')
        colon_count = 0
        total_lines_checked = min(20, len(lines))
        
        for i in range(total_lines_checked):
            line = lines[i]
            if ':' in line and line.count(':') >= 5:  # 进口舱单通常有很多冒号
                colon_count += 1
        
        if colon_count >= 3:  # 如果前20行中有3行以上有多个冒号，很可能是进口舱单
            logging.info(f"❌ 检测到进口舱单格式: 有{colon_count}行使用冒号分隔")
            return False
        
        # 特征4: 检查是否有51行和53行配对的结构
        has_51_line = False
        has_53_line = False
        
        for line in lines[:20]:  # 检查前20行
            if line.startswith('51') and len(line) >= 13:
                # 检查51行是否包含冒号（进口舱单特征）
                if ':' not in line:
                    has_51_line = True
            elif line.startswith('53') and len(line) >= 43:
                has_53_line = True
        
        # 如果同时有51行和53行，且51行没有冒号，很可能是出口舱单
        if has_51_line and has_53_line:
            logging.info("✅ 检测到出口舱单格式: 有51行和53行配对")
            return True
        elif has_51_line:
            logging.info("⚠️ 检测到可能有51行，但无53行")
            return True  # 还是尝试处理，可能是简化格式
        
        logging.info("❌ 未识别为出口舱单格式")
        return False
        
    except Exception as e:
        logging.error(f"判断舱单类型时出错: {e}")
        return False

def parse_txt_content(txt_content):
    """解析TXT文件内容，提取箱号、英文货名、中文货名和提单号信息"""
    try:
        # 首先检查是否为出口舱单
        if not is_export_manifest(txt_content):
            logging.warning("⚠️ 检测到非出口舱单格式，跳过处理")
            return None
        
        # 查找所有以51和53开头的记录行
        lines = txt_content.split('\n')
        
        # 存储所有记录
        all_records = []
        
        for line in lines:
            if line.startswith('51') or line.startswith('53'):
                all_records.append(line)
        
        if not all_records:
            logging.warning("未找到51或53记录行")
            return None
        
        logging.info(f"找到 {len(all_records)} 条记录")
        
        # 按顺序处理记录，不进行合并
        container_data = []
        matched_count = 0
        
        # 遍历所有记录
        for i, record in enumerate(all_records):
            if record.startswith('51') and len(record) >= 44:
                # 提取箱号（位置3-13，索引2:13）
                container_no = record[2:13].strip()
                
                # 提取提单号（位置29-44，索引28:44）- 这是DOCUMENT NO.
                bill_of_lading = ""
                if len(record) >= 44:
                    bill_of_lading = record[28:44].strip()
                    # 清理可能的空格
                    bill_of_lading = bill_of_lading.replace('\x00', '').strip()
                
                # 查找下一个53记录作为货名
                english_goods_description = "未知货名"
                for j in range(i+1, len(all_records)):
                    if all_records[j].startswith('53') and len(all_records[j]) >= 43:
                        # 使用下一个53记录作为货名，不检查箱号是否匹配
                        english_goods_description = all_records[j][13:43].strip()
                        break
                
                # 检查货名是否包含关键词
                found_keywords = check_keywords_in_text(english_goods_description)
                if found_keywords:
                    # 获取中文货名
                    chinese_goods_description = "未知中文货名"
                    if found_keywords:
                        # 如果有多个关键词，只取第一个进行翻译
                        main_keyword = found_keywords[0]
                        chinese_goods_description = get_chinese_goods_name(main_keyword, english_goods_description)
                    
                    container_data.append({
                        'container_no': container_no,
                        'english_goods_description': english_goods_description,
                        'chinese_goods_description': chinese_goods_description,
                        'bill_of_lading': bill_of_lading if bill_of_lading else "未知提单号"
                    })
                    matched_count += 1
                    logging.info(f"匹配到关键词 - 箱号: {container_no}, 提单号: {bill_of_lading}, 英文货名: {english_goods_description}, 中文货名: {chinese_goods_description}")
                else:
                    logging.info(f"未匹配关键词 - 箱号: {container_no}, 提单号: {bill_of_lading}, 英文货名: {english_goods_description}")
        
        if not container_data:
            logging.warning("未找到包含关键词的记录")
            return None
            
        logging.info(f"成功解析 {len(container_data)} 条匹配关键词的数据")
        return container_data
    except Exception as e:
        logging.error(f"解析TXT内容时出错: {e}")
        return None

def create_excel_file(container_data, excel_filename):
    """根据解析的数据创建Excel文件 - 四列版本：提单号、箱号、英文货名、中文货名"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "出口舱单"
        
        # 设置表头 - 四列
        ws['A1'] = '提单号'
        ws['B1'] = '箱号'
        ws['C1'] = '英文货名'
        ws['D1'] = '中文货名'
        
        # 填充数据
        for idx, data in enumerate(container_data, start=2):
            ws[f'A{idx}'] = data.get('bill_of_lading', '未知提单号')
            ws[f'B{idx}'] = data['container_no']
            ws[f'C{idx}'] = data['english_goods_description']
            ws[f'D{idx}'] = data['chinese_goods_description']
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25  # 提单号列宽
        ws.column_dimensions['B'].width = 20  # 箱号列宽
        ws.column_dimensions['C'].width = 40  # 英文货名列宽
        ws.column_dimensions['D'].width = 40  # 中文货名列宽
        
        # 保存Excel文件
        wb.save(excel_filename)
        
        # 验证文件是否创建成功
        if os.path.exists(excel_filename):
            file_size = os.path.getsize(excel_filename)
            logging.info(f"✅ Excel文件创建成功: {excel_filename}, 大小: {file_size} 字节")
            return True
        else:
            logging.error(f"❌ Excel文件未创建成功: {excel_filename}")
            return False
            
    except Exception as e:
        logging.error(f"❌ 创建Excel文件时出错: {e}")
        return False
    

def send_reply_with_attachment_fixed(to_addr, subject, excel_file_path, original_subject, email_type='export'):
    """发送回复邮件并附加Excel文件"""
    try:
        # 检查文件是否存在
        if not os.path.exists(excel_file_path):
            logging.error(f"❌ Excel文件不存在: {excel_file_path}")
            return False
        
        file_size = os.path.getsize(excel_file_path)
        logging.info(f"📊 Excel文件大小: {file_size} 字节")
        
        if file_size == 0:
            logging.error("❌ Excel文件为空")
            return False
        
        # 提取邮箱地址
        to_addr_clean = extract_email_address(to_addr)
        logging.info(f"📤 准备发送邮件到: {to_addr_clean} (原始: {to_addr})")
        logging.info(f"🔧 使用SMTP服务器: {smtp_server}:{smtp_port}")
        logging.info(f"📧 发件人: {email_address}")
        
        # 解码原邮件主题
        original_subject_decoded = decode_email_header(original_subject)
        # 获取Excel文件名（去除路径）
        excel_filename_only = os.path.basename(excel_file_path)
        
        # 获取配置的额外收件人
        additional_recipients = config_manager.get_additional_recipients(email_type)
        
        # 构建收件人列表：原始发件人 + 额外收件人
        recipients = [to_addr_clean]  # 主要收件人（提取后的邮箱地址）
        if additional_recipients:
            # 提取额外收件人的邮箱地址
            for addr in additional_recipients:
                clean_addr = extract_email_address(addr)
                if clean_addr and clean_addr != to_addr_clean:  # 避免重复
                    recipients.append(clean_addr)
        
        logging.info(f"📨 收件人列表: {recipients}")
        
        # 创建邮件 - 使用MIME格式
        msg = MIMEMultipart()
        msg['Subject'] = f'回复+{original_subject_decoded}+附件核查清单'
        msg['From'] = email_address
        msg['To'] = ', '.join(recipients)  # 设置所有收件人
        
        # 更新邮件正文内容
        body = f"""尊敬的客户，您好：

根据您发送至{email_address}邮箱、文件名为"{excel_filename_only}"的附件，经核查发现，相关航次中有涉及危险品（化学品）的货物。请贵司务必重视此事，并尽快采取应急处置措施。

如您需要其它协助，请随时与我们联系。

祝商祺！

天津港集装箱码头有限公司
危险品应急处置小组
"""
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # 添加Excel附件
        try:
            with open(excel_file_path, 'rb') as f:
                excel_attachment = MIMEApplication(
                    f.read(), 
                    _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                excel_attachment.add_header(
                    'Content-Disposition', 
                    'attachment', 
                    filename=os.path.basename(excel_file_path)
                )
                excel_attachment.add_header(
                    'Content-Type',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                msg.attach(excel_attachment)
            logging.info("✅ 附件添加成功")
        except Exception as e:
            logging.error(f"❌ 读取Excel文件失败: {e}")
            return False
        
        # 发送邮件
        logging.info("正在连接SMTP服务器...")
        try:
            server = smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=30)
            logging.info("✅ SMTP_SSL连接成功")
            
            server.ehlo()
            logging.info(f"✅ 服务器特性: {server.esmtp_features}")
            
        except Exception as e:
            logging.error(f"❌ SMTP连接失败: {e}")
            return False
        
        try:
            logging.info("🔐 正在登录...")
            server.login(email_address, password)
            logging.info("✅ 登录成功")
            
        except smtplib.SMTPAuthenticationError as e:
            logging.error(f"❌ 认证失败: {e}")
            server.quit()
            return False
        except Exception as e:
            logging.error(f"❌ 登录过程中出错: {e}")
            server.quit()
            return False
        
        try:
            logging.info("📨 正在发送邮件...")
            server.sendmail(email_address, recipients, msg.as_string())
            logging.info("✅ 邮件发送成功")
            
        except Exception as e:
            logging.error(f"❌ 发送失败: {e}")
            server.quit()
            return False
        
        server.quit()
        logging.info("🔌 连接已关闭")
        
        logging.info(f"✅ 回复邮件发送成功，收件人: {recipients}")
        
        # 发送成功后删除临时Excel文件
        try:
            os.remove(excel_file_path)
            logging.info(f"🗑️ 已删除临时文件: {excel_file_path}")
        except Exception as e:
            logging.warning(f"⚠️ 删除临时文件失败: {e}")
        
        return True
        
    except Exception as e:
        logging.error(f"❌ 发送回复邮件时出错: {e}")
        return False

def send_sms_notification(account, password, mobiles, content, error_info=""):
    """发送短信通知"""
    try:
        # 构建完整的短信内容
        full_content = f"{content}"
        if error_info:
            # 截断错误信息，避免短信过长
            error_short = error_info[:50] + "..." if len(error_info) > 50 else error_info
            full_content += f" 错误: {error_short}"
        
        logging.info(f"📱 准备发送短信通知到: {mobiles}")
        
        paras = {
            "action": "send",
            "account": account,
            "password": password,
            "mobile": mobiles,
            "content": full_content              
        }
        postdata = urllib.parse.urlencode(paras)
        
        # 创建请求
        req = urllib.request.Request(
            url=SMS_API_URL, 
            data=postdata.encode('utf-8'), 
            method='POST'
        )
        
        # 发送请求
        res = urllib.request.urlopen(req, timeout=10)
        response = res.read().decode()
        
        # 解析响应
        if response:
            doc = parseString(response)
            root = doc.documentElement
            
            returnstatus = root.getElementsByTagName("returnstatus")[0].childNodes[0].data
            message = root.getElementsByTagName("message")[0].childNodes[0].data
            
            if returnstatus == "Success":
                logging.info(f"✅ 短信发送成功: {message}")
                return True
            else:
                logging.error(f"❌ 短信发送失败: {message}")
                return False
        else:
            logging.error("❌ 短信发送返回空响应")
            return False
            
    except urllib.error.HTTPError as e:
        logging.error(f"❌ 短信发送HTTP错误: {e.code}, {e.reason}")
        return False
    except urllib.error.URLError as e:
        logging.error(f"❌ 短信发送URL错误: {e.reason}")
        return False
    except Exception as e:
        logging.error(f"❌ 短信发送失败: {e}")
        return False

def send_exit_notification(error_info="", is_manual=False):
    """发送程序退出通知"""
    if not SMS_ACCOUNT or not SMS_PASSWORD or not SMS_MOBILES:
        logging.warning("⚠️ 短信配置不完整，跳过短信通知")
        return False
    
    if is_manual:
        content = "【天津港集装箱码头有限公司】出口舱单处理程序已手动关闭"
    else:
        content = SMS_CONTENT_TEMPLATE
    
    return send_sms_notification(SMS_ACCOUNT, SMS_PASSWORD, SMS_MOBILES, content, error_info)

def process_email(msg, email_uid):
    """处理单封邮件"""
    try:
        # 获取邮件基本信息（解码邮件头）
        subject = decode_email_header(msg.get('subject', '无主题'))
        from_header = decode_email_header(msg.get('from', '未知发件人'))
        from_addr = extract_email_address(from_header)
        date = decode_email_header(msg.get('date', '未知日期'))
        
        logging.info(f"📧 处理邮件 - UID: {email_uid}, 发件人: {from_header}, 主题: {subject}")
        
        # 提取邮件正文
        email_body = get_email_body(msg)
        
        # 检查邮件主题和正文中的关键词
        found_keywords_in_subject = check_keywords_in_text(subject)
        found_keywords_in_body = check_keywords_in_text(email_body)
        
        # 收集所有附件文件名
        attachment_filenames = []
        found_keywords_in_attachments = []
        txt_attachments = []  # 存储TXT附件信息
        
        for part in msg.walk():
            content_disposition = str(part.get("Content-Disposition"))
            
            if "attachment" in content_disposition:
                filename = part.get_filename()
                if filename:
                    # 解码附件文件名
                    decoded_filename = decode_email_header(filename)
                    attachment_filenames.append(decoded_filename)
                    
                    # 检查附件文件名中的关键词
                    found_in_filename = check_keywords_in_text(decoded_filename)
                    if found_in_filename:
                        found_keywords_in_attachments.extend(found_in_filename)
                    
                    # 如果是TXT文件，保存附件内容
                    if decoded_filename.lower().endswith('.txt'):
                        try:
                            file_content = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                            
                            # 首先检查是否为出口舱单
                            if is_export_manifest(file_content):
                                txt_attachments.append({
                                    'filename': decoded_filename,
                                    'content': file_content
                                })
                                logging.info(f"📄 发现出口舱单TXT附件: {decoded_filename}")
                            else:
                                logging.info(f"📄 跳过非出口舱单TXT附件: {decoded_filename}")
                        except Exception as e:
                            logging.error(f"❌ 读取TXT附件 {decoded_filename} 时出错: {e}")
        
        # 处理TXT附件
        container_data = None
        matched_keywords_str = ""
        txt_attachment_name = ""
        container_count = 0
        excel_sent = 0
        english_goods_list = []
        chinese_goods_list = []
        
        if txt_attachments:
            for txt_attachment in txt_attachments:
                logging.info(f"🔍 开始解析TXT附件: {txt_attachment['filename']}")
                
                # 解析TXT内容
                container_data = parse_txt_content(txt_attachment['content'])
                
                if container_data:
                    container_count = len(container_data)
                    txt_attachment_name = txt_attachment['filename']
                    logging.info(f"✅ TXT附件解析成功，找到 {container_count} 条匹配关键词的数据")
                    
                    # 从所有箱子中收集中英文货名
                    english_goods_list = [item['english_goods_description'] for item in container_data]
                    chinese_goods_list = [item['chinese_goods_description'] for item in container_data]
                    
                    # 从所有箱子的货名中提取关键词
                    all_goods_descriptions = " ".join(english_goods_list)
                    matched_keywords = check_keywords_in_text(all_goods_descriptions)
                    matched_keywords_str = ",".join(matched_keywords) if matched_keywords else "出口舱单匹配"
                    
                    # 在这里添加统计记录
                    try:
                        add_attachment_statistics(
                            email_uid=email_uid,
                            sender=from_header,
                            sender_address=from_addr,
                            subject=subject,
                            received_date=date,
                            txt_attachment=txt_attachment_name,
                            container_data=container_data,
                            matched_keywords=matched_keywords_str
                        )
                    except Exception as e:
                        logging.warning(f"⚠️ 添加统计记录时出错: {e}，但继续处理邮件")
                    
                    # 创建Excel文件
                    base_name = os.path.splitext(txt_attachment['filename'])[0]
                    excel_filename = f"processed_{base_name}.xlsx"
                    
                    if create_excel_file(container_data, excel_filename):
                        # 发送回复邮件
                        # 发送回复邮件
                        if send_reply_with_attachment_fixed(from_header, subject, excel_filename, subject, 'export'):
                            excel_sent = 1
                            logging.info(f"✅ 完整处理流程成功，匹配关键词: {matched_keywords_str}")
                            
                            # 保存到数据库
                            attachment_names_str = ",".join(attachment_filenames) if attachment_filenames else ""
                            english_goods_str = ",".join(english_goods_list)
                            chinese_goods_str = ",".join(chinese_goods_list)
                            
                            save_keyword_email(
                                email_uid=email_uid,
                                sender=from_header,
                                sender_address=from_addr,
                                subject=subject,
                                received_date=date,
                                matched_keywords=matched_keywords_str,
                                txt_attachment=txt_attachment_name,
                                container_count=container_count,
                                attachment_names=attachment_names_str,
                                english_goods_descriptions=english_goods_str,
                                chinese_goods_descriptions=chinese_goods_str
                            )
                        else:
                            logging.error("❌ 发送回复邮件失败")
                            try:
                                if os.path.exists(excel_filename):
                                    os.remove(excel_filename)
                                    logging.info(f"🗑️ 已清理临时文件: {excel_filename}")
                            except Exception as e:
                                logging.warning(f"⚠️ 清理临时文件失败: {e}")
                    else:
                        logging.error("❌ 创建Excel文件失败")
                else:
                    logging.warning("⚠️ 非指定格式的TXT文件无法转化或未找到关键词匹配")
        else:
            logging.info("📭 未发现出口舱单TXT附件")
        
        # 合并所有找到的关键词
        all_found_keywords = found_keywords_in_subject + found_keywords_in_body + found_keywords_in_attachments
        
        # 记录邮件处理状态到日志
        has_keyword = len(all_found_keywords) > 0 or container_data is not None
        
        if has_keyword:
            # 去重
            unique_keywords = list(set(all_found_keywords))
            if container_data and not matched_keywords_str:
                # 如果是从TXT附件解析到的，但没有提取到具体关键词，使用通用描述
                matched_keywords_str = "TXT附件匹配"
            elif not matched_keywords_str and unique_keywords:
                matched_keywords_str = ",".join(unique_keywords)
            
            logging.info(f"🎯 检测到关键词匹配: {matched_keywords_str}")
            
            # 记录到日志文件
            log_email_processed(
                email_uid=email_uid,
                sender=from_header,
                subject=subject,
                has_keyword=True,
                excel_sent=excel_sent,
                matched_keywords=matched_keywords_str,
                container_count=container_count
            )
            
            return True, from_addr, subject, "keyword_match", matched_keywords_str, excel_sent
        else:
            # 记录到日志文件（没有关键词）
            log_email_processed(
                email_uid=email_uid,
                sender=from_header,
                subject=subject,
                has_keyword=False,
                excel_sent=0,
                matched_keywords="",
                container_count=0
            )
            
            logging.info(f"📭 邮件未匹配关键词 - 主题: {subject}, 发件人: {from_header}")
            return False, None, None, "no_keyword", "", 0
        
    except Exception as e:
        logging.error(f"❌ 处理邮件时出错: {e}")
        return False, None, None, None, "", 0

def get_email_uids(server):
    """安全地获取所有邮件的UID列表"""
    try:
        # 方法1：使用uidl命令获取所有UID
        response, uid_list, _ = server.uidl()
        uids = []
        for uid_line in uid_list:
            # 将字节转换为字符串并提取UID
            uid_str = uid_line.decode('utf-8')
            # 格式通常是 "序号 UID"，我们只需要UID部分
            parts = uid_str.split()
            if len(parts) >= 2:
                uids.append(parts[1])
        return uids
    except Exception as e:
        logging.error(f"❌ 获取UID列表时出错: {e}")
        # 如果上面的方法失败，尝试逐封邮件获取UID
        try:
            email_count, _ = server.stat()
            uids = []
            for i in range(1, email_count + 1):
                # 使用更安全的方式获取UID
                result = server.uidl(i)
                # 处理不同格式的返回结果
                if len(result) == 2:
                    # 有些服务器返回 (response, data)
                    _, uid_data = result
                else:
                    # 标准格式 (response, data, octets)
                    _, uid_data, _ = result
                
                # 提取UID
                uid_str = uid_data.decode('utf-8').split()[-1]
                uids.append(uid_str)
            return uids
        except Exception as e2:
            logging.error(f"❌ 备用方法获取UID列表也失败: {e2}")
            return []

def get_email_received_datetime(server, msg_no: int):
    """尽量只获取邮件头部并解析 Date，用于“最近N天扫描”优化"""
    try:
        # POP3 TOP 0: 只取头部，不取正文，速度快
        resp, lines, _ = server.top(msg_no, 0)
        raw = b'\r\n'.join(lines).decode('utf-8', errors='ignore')
        msg = Parser(policy=default).parsestr(raw)
        date_hdr = msg.get('Date')
        if not date_hdr:
            return None
        dt = parsedate_to_datetime(date_hdr)
        # 统一成 naive datetime（本地时间）用于比较
        if getattr(dt, 'tzinfo', None) is not None:
            dt = dt.astimezone().replace(tzinfo=None)
        return dt
    except Exception:
        return None

def get_keyword_emails_count():
    """获取关键词邮件数量（数据库中）"""
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM keyword_emails')
        count = cursor.fetchone()[0]
        conn.close()
        return count
    except Exception as e:
        logging.error(f"❌ 获取关键词邮件数量失败: {e}")
        return 0

def get_today_keyword_emails():
    """获取今天的关键词邮件数量（数据库中）"""
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        today = datetime.now().strftime('%Y-%m-%d')
        cursor.execute('SELECT COUNT(*) FROM keyword_emails WHERE DATE(processed_date) = ?', (today,))
        count = cursor.fetchone()[0]
        conn.close()
        return count
    except Exception as e:
        logging.error(f"❌ 获取今天关键词邮件数量失败: {e}")
        return 0

def view_log_summary():
    """查看日志文件摘要"""
    try:
        if not os.path.exists(LOG_CSV_FILE):
            print(f"❌ 日志文件不存在: {LOG_CSV_FILE}")
            return
        
        print("📊 邮件处理日志摘要")
        print("=" * 60)
        
        with open(LOG_CSV_FILE, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader)  # 读取标题行
            
            # 统计信息
            total_count = 0
            keyword_count = 0
            excel_sent_count = 0
            today = datetime.now().strftime('%Y-%m-%d')
            today_count = 0
            
            for row in reader:
                total_count += 1
                if len(row) >= 5 and row[4] == '1':  # has_keyword列
                    keyword_count += 1
                if len(row) >= 6 and row[5] == '1':  # excel_sent列
                    excel_sent_count += 1
                if len(row) >= 1 and row[0].startswith(today):  # timestamp列
                    today_count += 1
        
        print(f"📈 统计信息:")
        print(f"   总处理邮件: {total_count} 封")
        print(f"   关键词邮件: {keyword_count} 封")
        print(f"   已发送Excel: {excel_sent_count} 封")
        print(f"   今日处理邮件: {today_count} 封")
        
        # 显示最新记录
        print(f"\n📨 最新处理记录 (最近5条):")
        
        # 重新读取文件获取最后几行
        with open(LOG_CSV_FILE, 'r', newline='', encoding='utf-8') as f:
            lines = f.readlines()
        
        # 显示最后5条记录（不包括标题行）
        for line in lines[-5:]:
            if line.startswith('timestamp'):
                continue  # 跳过标题行
            
            parts = line.strip().split(',')
            if len(parts) >= 4:
                timestamp = parts[0]
                uid = parts[1][:10] + '...' if len(parts[1]) > 10 else parts[1]
                sender = parts[2][:20] + '...' if len(parts[2]) > 20 else parts[2]
                subject = parts[3][:30] + '...' if len(parts[3]) > 30 else parts[3]
                
                print(f"   {timestamp} | {uid} | {sender} | {subject}")
        
    except Exception as e:
        print(f"❌ 查看日志摘要失败: {e}")

def view_database_simple():
    """简单的数据库查看函数"""
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        
        print("📊 关键词邮件数据库（仅包含已发送Excel的邮件）")
        print("=" * 60)
        
        # 获取统计信息
        cursor.execute('SELECT COUNT(*) FROM keyword_emails')
        total = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM keyword_emails WHERE DATE(processed_date) = DATE("now")')
        today = cursor.fetchone()[0]
        
        print(f"📈 统计信息:")
        print(f"   总关键词邮件: {total} 封")
        print(f"   今日关键词邮件: {today} 封")
        
        # 关键词统计
        print(f"\n🔑 关键词匹配统计:")
        cursor.execute('SELECT matched_keywords, COUNT(*) as count FROM keyword_emails GROUP BY matched_keywords ORDER BY count DESC')
        keyword_stats = cursor.fetchall()
        
        for keyword, count in keyword_stats:
            print(f"   {keyword}: {count} 次")
        
        # 显示最新记录
        print(f"\n📨 最新关键词邮件 (最近10条):")
        query = '''
            SELECT 
                id,
                strftime('%Y-%m-%d %H:%M', processed_date) as 处理时间,
                sender as 发件人,
                subject as 主题,
                matched_keywords as 匹配关键词,
                container_count as 箱子数量
            FROM keyword_emails 
            ORDER BY processed_date DESC 
            LIMIT 10
        '''
        
        cursor.execute(query)
        rows = cursor.fetchall()
        if rows:
            print(f"{'ID':<5} {'时间':<20} {'发件人':<30} {'主题':<50} {'关键词':<20} {'箱子':<10}")
            print("-" * 135)
            for row in rows:
                # 截断过长的字段
                sender_display = (row[2][:28] + '...') if len(row[2]) > 30 else row[2]
                subject_display = (row[3][:48] + '...') if len(row[3]) > 50 else row[3]
                keywords_display = (row[4][:18] + '...') if len(row[4]) > 20 else row[4]
                
                print(f"{row[0]:<5} {row[1]:<20} {sender_display:<30} {subject_display:<50} {keywords_display:<20} {row[5]:<10}")
        else:
            print("   暂无数据")
        
        conn.close()
        
    except Exception as e:
        print(f"❌ 查看数据库失败: {e}")

def main():
    logging.info("🚀 启动邮件自动处理程序...")
    logging.info(f"📧 邮箱: {email_address}")
    logging.info(f"🔑 关键词: {keywords}")
    logging.info(f"📮 SMTP服务器: {smtp_server}:{smtp_port}")
    logging.info(f"🗄️  关键词邮件数据库: {db_file}")
    logging.info(f"📝 邮件处理日志: {LOG_CSV_FILE}")
    
    # 初始化日志文件
    if not init_log_file():
        logging.error("❌ 日志文件初始化失败，程序退出")
        return
    
    # 初始化数据库
    if not init_database():
        logging.error("❌ 数据库初始化失败，程序退出")
        return
    
    # 显示统计信息
    keyword_count = get_keyword_emails_count()
    today_keyword = get_today_keyword_emails()
    
    logging.info(f"📊 数据库统计 - 总关键词邮件: {keyword_count} 封, 今日: {today_keyword} 封")
    
    check_interval = 30  # 默认检查间隔为30秒
    
    # 添加一个计数器，每24小时清理一次旧日志
    cleanup_counter = 0
    cleanup_interval = 24 * 60 * 60  # 24小时（秒）
    
    try:
        while True:
            try:
                current_time = time.strftime('%Y-%m-%d %H:%M:%S')
                logging.info(f"⏰ {current_time} 开始检查新邮件...")
                
                # 定期清理日志文件（每24小时一次）
                cleanup_counter += check_interval
                if cleanup_counter >= cleanup_interval:
                    cleanup_old_log_entries()
                    cleanup_counter = 0
                
                # 连接POP3服务器
                logging.info(f"🔗 正在连接服务器 {pop3_server}:{pop3_port}...")
                server = poplib.POP3_SSL(pop3_server, pop3_port, timeout=30)
                logging.info("✅ 服务器连接成功！")
                
                # 登录邮箱
                logging.info("🔐 正在登录邮箱...")
                server.user(email_address)
                server.pass_(password)
                logging.info("✅ 邮箱登录成功！")
                
                # 获取邮件统计信息
                email_count, total_size = server.stat()
                logging.info(f"📬 邮箱中共有 {email_count} 封邮件")
                
                # 获取所有邮件的UID
                all_uids = get_email_uids(server)
                if all_uids:
                    logging.info(f"📋 成功获取 {len(all_uids)} 个邮件UID")
                else:
                    logging.warning("⚠️ 未能获取邮件UID列表，跳过本次检查")
                    server.quit()
                    time.sleep(check_interval)
                    continue
                
                # 处理新邮件
                new_emails_processed = 0
                keyword_emails_found = 0
                
                cutoff_scan_time = datetime.now() - timedelta(days=SCAN_DAYS)

                # POP3 的序号通常按时间从旧到新排列：1最旧，N最新。
                # 这里从最新开始逆序处理，遇到超过 SCAN_DAYS 的邮件则直接停止遍历。
                for i in range(min(email_count, len(all_uids)), 0, -1):
                    try:
                        uid = all_uids[i-1]
                        
                        # 检查邮件是否已在日志中处理过
                        if is_email_processed(uid):
                            continue
                        
                        # “最近N天”过滤：尽量只取头部判断日期
                        received_dt = get_email_received_datetime(server, i)
                        if received_dt and received_dt < cutoff_scan_time:
                            # 该邮件已早于扫描窗口；因为在倒序遍历，后续只会更旧，直接停止。
                            break

                        # 安全地获取邮件内容
                        try:
                            result = server.retr(i)
                            if len(result) >= 2:
                                lines = result[1]
                                msg_content = b'\r\n'.join(lines).decode('utf-8', errors='ignore')
                                msg = Parser(policy=default).parsestr(msg_content)
                                
                                # 处理邮件
                                has_match, from_addr, subject, match_type, matched_keywords, excel_sent = process_email(msg, uid)
                                
                                if has_match:
                                    keyword_emails_found += 1
                                
                                new_emails_processed += 1
                                
                                # 处理完一封邮件后稍作休息，避免服务器压力
                                time.sleep(0.5)
                            else:
                                logging.error(f"❌ 获取第 {i} 封邮件内容失败")
                        except Exception as e:
                            logging.error(f"❌ 处理第 {i} 封邮件内容时出错: {e}")
                            continue
                        
                    except Exception as e:
                        logging.error(f"❌ 处理第 {i} 封邮件时出错: {e}")
                        continue
                
                if new_emails_processed > 0:
                    if keyword_emails_found > 0:
                        logging.info(f"✅ 本轮处理完成，共处理 {new_emails_processed} 封新邮件，发现 {keyword_emails_found} 封关键词邮件")
                        check_interval = 30  # 30秒后再次检查
                    else:
                        logging.info(f"📭 本轮处理完成，共处理 {new_emails_processed} 封新邮件，未发现关键词邮件")
                        check_interval = 30  # 30秒后再次检查
                else:
                    logging.info("📭 没有发现新邮件需要处理")
                    check_interval = 60  # 60秒后再次检查
                
                # 关闭连接
                server.quit()
                logging.info("🔌 已断开服务器连接")
                
                # 更新统计信息
                today_keyword = get_today_keyword_emails()
                logging.info(f"📊 更新统计 - 今日关键词邮件: {today_keyword} 封")
                
            except poplib.error_proto as e:
                error_msg = f"POP3协议错误: {e}"
                logging.error(f"❌ {error_msg}")
                if "Unable to log on" in str(e) or "Authentication failed" in str(e):
                    logging.error("🔐 登录失败，请检查邮箱地址和密码/授权码是否正确")
                    # 发送短信通知
                    send_exit_notification(f"邮箱登录失败: {str(e)[:50]}")
                    check_interval = 30
            except Exception as e:
                error_msg = f"发生错误: {e}"
                logging.error(f"❌ {error_msg}")
                # 发送短信通知
                send_exit_notification(str(e)[:100])
                check_interval = 30
            
            # 等待一段时间后再次检查
            logging.info(f"⏳ 等待{check_interval}秒后再次检查...")
            time.sleep(check_interval)
            
    except Exception as e:
        # 捕获主循环外的异常
        error_msg = f"主程序异常: {e}"
        logging.error(f"❌ {error_msg}")
        send_exit_notification(str(e)[:100])
        raise  # 重新抛出异常

if __name__ == "__main__":
    try:
        # 添加命令行参数支持
        import sys
        if len(sys.argv) > 1:
            if sys.argv[1] == 'view':
                # 查看数据库
                view_database_simple()
            elif sys.argv[1] == 'log':
                # 查看日志摘要
                view_log_summary()
            elif sys.argv[1] == 'test_sms':  # 添加测试命令
                print("🧪 测试短信通知功能...")
                # 测试异常退出通知
                send_exit_notification("这是测试异常信息", is_manual=False)
                print("已发送异常退出测试短信")
                
                # 测试手动关闭通知
                send_exit_notification(is_manual=True)
                print("已发送手动关闭测试短信")
                
                print("✅ 短信测试完成，请检查手机是否收到短信")
            else:
                print("使用方法:")
                print("  python OutputAutoRWwithSend_3_0.py           # 运行主程序")
                print("  python OutputAutoRWwithSend_3_0.py view      # 查看数据库")
                print("  python OutputAutoRWwithSend_3_0.py log       # 查看处理日志")
                print("  python OutputAutoRWwithSend_3_0.py test_sms  # 测试短信通知")
        else:
            # 运行主程序
            main()
    except KeyboardInterrupt:
        logging.info("👋 程序被用户中断")
        # 发送手动关闭通知
        send_exit_notification(is_manual=True)
    except Exception as e:
        logging.error(f"❌ 程序异常退出: {e}")
        # 发送异常退出通知
        send_exit_notification(str(e)[:100])